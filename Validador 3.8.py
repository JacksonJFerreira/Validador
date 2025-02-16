import customtkinter as ctk
from tkinter import filedialog, ttk, messagebox
import openpyxl
from openpyxl.styles import Font, Border, Side
import os
from datetime import datetime
import logging

class ValidadorDeDados:
    def __init__(self):
        self.setup_logging()
        self.setup_gui()

    def setup_logging(self):
        logging.basicConfig(filename='validador.log', level=logging.INFO,
                            format='%(asctime)s - %(levelname)s - %(message)s')

    def setup_gui(self):
        ctk.set_appearance_mode("Sytsem")
        ctk.set_default_color_theme("blue")

        self.janela = ctk.CTk()
        self.janela.title("Validação Nominal de Empregados")
        self.janela.geometry("500x350")
        self.janela.resizable(False, False)  # Trava o redimensionamento

        self.diretorio_sifac = ctk.StringVar()
        self.diretorio_sispat = ctk.StringVar()
        self.mensagem_status = ctk.StringVar()

        self.criar_widgets()

    def criar_widgets(self):
        rotulo_titulo = ctk.CTkLabel(
            self.janela, text="Selecionar Pastas de Arquivos Excel", font=("Aptos", 14))
        rotulo_titulo.pack(pady=10)

        entrada_sifac = ctk.CTkEntry(self.janela, textvariable=self.diretorio_sifac, width=400)
        entrada_sifac.pack(pady=5)
        botao_selecionar_sifac = ctk.CTkButton(
            self.janela, text="Selecionar Pasta SIFAC", command=self.selecionar_pasta_sifac)
        botao_selecionar_sifac.pack(pady=5)

        entrada_sispat = ctk.CTkEntry(self.janela, textvariable=self.diretorio_sispat, width=400)
        entrada_sispat.pack(pady=5)
        botao_selecionar_sispat = ctk.CTkButton(
            self.janela, text="Selecionar Pasta SISPAT", command=self.selecionar_pasta_sispat)
        botao_selecionar_sispat.pack(pady=5)

        # Botões lado a lado
        botao_validar = ctk.CTkButton(self.janela, text="Validar Dados", command=self.validar_dados)
        botao_lista_sispat = ctk.CTkButton(self.janela, text="Lista SISPAT", command=self.criar_lista_sispat)
        
        botao_validar.place(relx=0.3, rely=0.62, anchor="center")
        botao_lista_sispat.place(relx=0.7, rely=0.62, anchor="center")

        self.progresso = ttk.Progressbar(
            self.janela, orient="horizontal", length=400, mode="determinate")
        self.progresso.pack(pady=60)
        
        rotulo_status = ctk.CTkLabel(
            self.janela, textvariable=self.mensagem_status, font=("Aptos", 12))
        rotulo_status.pack(pady=20)

    def selecionar_pasta_sifac(self):
        self.diretorio_sifac.set(filedialog.askdirectory())

    def selecionar_pasta_sispat(self):
        self.diretorio_sispat.set(filedialog.askdirectory())

    def obter_competencia_sispat(self, caminho):
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            sheet = wb.active
            competencia = sheet['E2'].value
            logging.info(f"Competência SISPAT ({caminho}): {competencia}")
            return competencia
        except Exception as e:
            logging.error(f"Erro ao ler competência SISPAT ({caminho}): {e}")
            return None

    def obter_competencia_sifac(self, caminho):
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            sheet = wb.active
            competencia = sheet['F5'].value
            if competencia is None:
                competencia = ''
            if isinstance(competencia, str):
                if "competencia:" in competencia.lower():
                    competencia = competencia.split(":")[-1].strip()
                else:
                    competencia = competencia.strip()
            logging.info(f"Competência SIFAC ({caminho}): {competencia}")
            return competencia
        except Exception as e:
            logging.error(f"Erro ao ler competência SIFAC ({caminho}): {e}")
            return None

    def formatar_competencia(self, competencia):
        if competencia is None:
            return None
        if isinstance(competencia, datetime):
            return competencia.strftime("%Y-%m")
        elif isinstance(competencia, str):
            competencia = competencia.strip()
            if '/' in competencia:
                mes, ano = competencia.split('/')
                return f"{ano}-{mes.zfill(2)}"
            try:
                return datetime.strptime(competencia, "%Y-%m-%d").strftime("%Y-%m")
            except ValueError:
                pass
            try:
                return datetime.strptime(competencia, "%Y/%m/%d").strftime("%Y-%m")
            except ValueError:
                pass
            if len(competencia) >= 6:
                return f"{competencia[:4]}-{competencia[4:6]}"
        return str(competencia).strip()

    def validar_competencia(self):
        try:
            arquivo_sifac = next(f for f in os.listdir(self.diretorio_sifac.get()) if f.endswith(".xlsx"))
            caminho_arquivo_sifac = os.path.join(self.diretorio_sifac.get(), arquivo_sifac)
            competencia_sifac = self.formatar_competencia(self.obter_competencia_sifac(caminho_arquivo_sifac))

            arquivo_sispat = next(f for f in os.listdir(self.diretorio_sispat.get()) if f.endswith(".xlsx"))
            caminho_arquivo_sispat = os.path.join(self.diretorio_sispat.get(), arquivo_sispat)
            competencia_sispat = self.formatar_competencia(self.obter_competencia_sispat(caminho_arquivo_sispat))

            if competencia_sifac != competencia_sispat:
                messagebox.showerror("Erro de Competência", "As competências não são iguais nos arquivos das pastas SIFAC e SISPAT.")
                return False
            return True
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao verificar competências: {e}")
    def criar_lista_sispat(self):
        try:
            arquivos_sifac = [f for f in os.listdir(self.diretorio_sifac.get()) if f.endswith(".xlsx")]
            arquivo_sispat = next(f for f in os.listdir(self.diretorio_sispat.get()) if f.endswith(".xlsx"))
        
            caminho_sispat = os.path.join(self.diretorio_sispat.get(), arquivo_sispat)
            wb_sispat = openpyxl.load_workbook(caminho_sispat)
            ws_sispat = wb_sispat["Relatorio_Empregado"]
        
            total_arquivos = len(arquivos_sifac)
            self.progresso["maximum"] = total_arquivos
        
            for i, arquivo_sifac in enumerate(arquivos_sifac):
                caminho_sifac = os.path.join(self.diretorio_sifac.get(), arquivo_sifac)
                wb_sifac = openpyxl.load_workbook(caminho_sifac)
            
                ws_sifac = wb_sifac["Relatorio_Empregado"]
                contrato_valor = str(ws_sifac['C5'].value).strip()
            
                if "Contrato:" in contrato_valor:
                    contrato_sifac = contrato_valor.split("Contrato:")[1].strip()
                else:
                    contrato_sifac = contrato_valor
            
                if not contrato_sifac:
                    continue
            
                # Obter lista de nomes encontrados na aba Dados_Validados
                nomes_encontrados = set()
                if "Dados_Validados" in wb_sifac.sheetnames:
                    ws_validados = wb_sifac["Dados_Validados"]
                    ultima_coluna = ws_validados.max_column
                
                    for row in ws_validados.iter_rows(min_row=6):
                        status = row[ultima_coluna - 2].value  # Coluna STATUS
                        if status == "Encontrado":
                            nome = row[3].value  # Coluna do nome
                            nomes_encontrados.add(nome)
            
                # Criar nova aba LISTA_SISPAT
                if "LISTA_SISPAT" in wb_sifac.sheetnames:
                    wb_sifac.remove(wb_sifac["LISTA_SISPAT"])
                ws_lista = wb_sifac.create_sheet("LISTA_SISPAT")
            
                ws_lista.append(["Contrato", "Nome"])
                for cell in ws_lista[1]:
                    cell.font = Font(bold=True)
            
                # Adicionar apenas nomes que não estão na lista de encontrados
                for row in ws_sispat.iter_rows(min_row=5, values_only=True):
                    if row[0] and str(row[0]).strip() == contrato_sifac:
                        if row[1] not in nomes_encontrados:
                            ws_lista.append([row[0], row[1]])
            
                self.ajustar_largura_colunas(ws_lista)
            
                self.progresso["value"] = i + 1
                self.janela.update_idletasks()
            
                wb_sifac.save(caminho_sifac)
        
            self.mensagem_status.set("Lista SISPAT criada com sucesso!")
        
        except Exception as e:


            self.mensagem_status.set(f"Erro ao criar lista SISPAT: {e}")
            logging.error(f"Erro ao criar lista SISPAT: {e}")
    def validar_dados(self):
        if not self.validar_competencia():
            return

        try:
            arquivos_sifac = [f for f in os.listdir(self.diretorio_sifac.get()) if f.endswith(".xlsx")]
            arquivos_sispat = [f for f in os.listdir(self.diretorio_sispat.get()) if f.endswith(".xlsx")]

            empregados_sispat = {}
            for arquivo in arquivos_sispat:
                caminho_arquivo = os.path.join(self.diretorio_sispat.get(), arquivo)
                wb = openpyxl.load_workbook(caminho_arquivo)
                ws = wb["Relatorio_Empregado"]
                for row in ws.iter_rows(min_row=5, min_col=2, max_col=2, values_only=True):
                    empregado = row[0]
                    if empregado:
                        empregados_sispat[empregado] = arquivo

            total_arquivos = len(arquivos_sifac)
            self.progresso["maximum"] = total_arquivos

            for i, arquivo in enumerate(arquivos_sifac):
                caminho_arquivo = os.path.join(self.diretorio_sifac.get(), arquivo)
                wb = openpyxl.load_workbook(caminho_arquivo)
                ws_origem = wb["Relatorio_Empregado"]

                if "Dados_Validados" in wb.sheetnames:
                    ws_destino = wb["Dados_Validados"]
                    wb.remove(ws_destino)
                ws_destino = wb.create_sheet("Dados_Validados")

                for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                    ws_destino.append(row)

                for row in ws_origem.iter_rows(min_row=4, values_only=True):
                    ws_destino.append(row)

                ultima_coluna = ws_destino.max_column

                ws_destino.cell(row=5, column=ultima_coluna + 1, value="STATUS")
                ws_destino.cell(row=5, column=ultima_coluna + 2, value="DATA E HORA")
                ws_destino.cell(row=5, column=ultima_coluna + 3, value="NOME DO ARQUIVO VALIDADO")

                for col in range(1, ultima_coluna + 4):
                    ws_destino.cell(row=5, column=col).font = Font(bold=True)

                coluna_status = ultima_coluna + 1
                coluna_data_hora = ultima_coluna + 2
                coluna_arquivo_validado = ultima_coluna + 3

                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                for idx, row in enumerate(ws_destino.iter_rows(min_row=6), start=6):
                    empregado = row[3].value
                    if empregado in empregados_sispat:
                        status = "Encontrado"
                        nome_arquivo_validado = empregados_sispat[empregado]
                    else:
                        status = "Não Encontrado"
                        nome_arquivo_validado = ""

                    ws_destino.cell(row=idx, column=coluna_status, value=status)
                    ws_destino.cell(row=idx, column=coluna_data_hora, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    ws_destino.cell(row=idx, column=coluna_arquivo_validado, value=nome_arquivo_validado)

                    for col in [coluna_status, coluna_data_hora, coluna_arquivo_validado]:
                        ws_destino.cell(row=idx, column=col).border = thin_border

                for cell in ws_destino.iter_rows(min_row=5, max_row=5):
                    for c in cell:
                        c.border = thin_border

                for row in ws_destino.iter_rows(min_row=1):
                    for cell in row:
                        cell.font = Font(name="Aptos Narrow", size=10)
                        cell.border = thin_border

                self.remover_colunas_em_branco(ws_destino)
                self.ajustar_largura_colunas(ws_destino)

                wb.save(caminho_arquivo)

                self.progresso["value"] = i + 1
                self.janela.update_idletasks()

            self.mensagem_status.set("Processamento concluído!")
        except Exception as e:
            self.mensagem_status.set(f"Erro ao processar: {e}")
            logging.error(f"Erro ao processar: {e}")

    def remover_colunas_em_branco(self, worksheet):
        colunas_a_remover = []
        for col in range(1, worksheet.max_column + 1):
            if all(worksheet.cell(row=row, column=col).value is None for row in range(1, worksheet.max_row + 1)):
                colunas_a_remover.append(col)

        for col in reversed(colunas_a_remover):
            worksheet.delete_cols(col)

    def ajustar_largura_colunas(self, worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    def run(self):
        self.janela.mainloop()

if __name__ == "__main__":
    app = ValidadorDeDados()
    app.run()
